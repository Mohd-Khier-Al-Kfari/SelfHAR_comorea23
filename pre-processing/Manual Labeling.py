# -*- coding: utf-8 -*-
"""
Created on Wed May 25 10:28:10 2022
Copy the Labels to the activity data
@author: Mohammad
"""
import math
from glob import glob
import os
import numpy as np
import pandas as pd
import datetime as dt
import openpyxl



"Read the labels"
def readXlsxFile(file):
    DataDic = {}
    sheets = openpyxl.load_workbook(file)
    for sheet in sheets.sheetnames:
        data = pd.read_excel(file, sheet_name=sheet)
        DataDic[sheet] = data
    return sheets.sheetnames, DataDic

"Prepare dataset"
def dataPrepare(DatasetFiles, User):
    dataFolder = [s for s in DatasetFiles if User in s]
    if dataFolder == []:
        dataset = pd.DataFrame()
        return dataset
    dataset = pd.read_csv(dataFolder[0])
    dataset.columns = ["Time", "acc_x", "acc_y", "acc_y"]
    dataset["Label"] = np.NaN
    dataset['Time'] = pd.to_datetime(dataset['Time'])
    dataset.sort_values(by="Time", inplace=True)
    dataset = dataset.reset_index(drop=True)
    return dataset

"Labeling the dataset"
def Labeling(dataset, ActivityDic, User):
    for j in range(len(ActivityDic[User])):
        print(User)
        Start = pd.to_datetime(ActivityDic[User]['Time'][j][0:ActivityDic[User]['Time'][j].find('-')-1]).replace(year=dataset['Time'][0].year,
                                                                                        month=dataset['Time'][0].month,
                                                                                        day=dataset['Time'][0].day)

        End = pd.to_datetime(ActivityDic[User]['Time'][j][ActivityDic[User]['Time'][j].find('-') + 2:]).replace(year=dataset['Time'][0].year,
                                                                                        month=dataset['Time'][0].month,
                                                                                        day=dataset['Time'][0].day)


        dataset.loc[(dataset["Time"] >=Start) & (dataset["Time"] <= End), "Label"] = ActivityDic[User]['Activity'][j]
    dataset = dataset.dropna(subset=['Label'])
    dataset = dataset.reset_index(drop=True)
    return dataset


"Location of the labels file"
LabelFolder = "/home/mkfari/Project/Labeled dataset/Daten Pilot_innen/Activities.xlsx"
#LabelFolder = "/Users/Mohammad/OneDrive/Desktop/Mechatronics MS/Uni/Sixth semester/Master Thesis/Labeled dataset/Daten Pilot_innen/Activities.xlsx"

"Location of the dataset"
#dataset_folder = "/Users/Mohammad/OneDrive/Desktop/Mechatronics MS/Uni/Sixth semester/Master Thesis//Labeled dataset/Daten Pilot_innen/"
DatasetFolder = "/home/mkfari/Project/Labeled dataset/decompressed/data"

"Location to save the labeled data"
LabeledDataFile = "/home/mkfari/Project/Labeled dataset/decompressed/Labeled Data/"


DatasetFiles = glob(DatasetFolder + "/*.csv") #Read names of dataset folders
Users, ActivityDic = readXlsxFile(LabelFolder) #Read labels file

for User in Users:
    #User = "P-3582_Tag7"
    dataset = dataPrepare(DatasetFiles= DatasetFiles, User=User) #Prepare and clean dataset
    if not dataset.empty:
        dataset = Labeling(dataset=dataset, ActivityDic=ActivityDic, User=User) #Dataset labeling
        dataset.to_csv(LabeledDataFile + User + ".csv") #Save dataset
